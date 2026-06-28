"""Render standard reports to a downloadable Excel workbook or CSV.

All renderers consume the same in-memory report data (the JSON shape the live
services return) via ``_section()`` so the file always matches what the UI shows.
``build_workbook`` (openpyxl) renders one sheet per report — used for both a single
report and the combined "Export all" pack. ``build_csv`` renders one report as a
single CSV table. Tabular by design — the standard ITSM export shape.
"""

from __future__ import annotations

import csv
import io

# Human titles + column layout for each report key. ``cols`` is a list of
# (header, row-key) pairs; reports that return a single summary dict are handled
# separately below.
REPORT_TITLES = {
    "ticket-data": "Ticket Data",
    "open-tickets": "Open tickets by project",
    "by-status": "Tickets by status",
    "by-priority": "Tickets by priority",
    "by-group": "Tickets by team",
    "agent-performance": "Agent performance",
    "sla-compliance": "SLA compliance",
    "resolution-trends": "Resolution trend (daily)",
    "volume-trends": "Ticket volume trend (daily)",
    "created-vs-resolved": "Created vs resolved",
    "resolution-time-by-priority": "Resolution time by priority",
    "sla-breach-list": "SLA breach list",
    "backlog-aging": "Backlog aging",
}

# (header label, key) columns for the list-shaped reports.
_TABLES = {
    "by-status": [("Status", "label"), ("Tickets", "value")],
    "by-priority": [("Priority", "label"), ("Tickets", "value")],
    "by-group": [("Team", "label"), ("Tickets", "value")],
    "agent-performance": [
        ("Agent", "agent"), ("Resolved", "resolved_count"),
        ("Open", "open_count"), ("Avg resolution (h)", "avg_resolution_hours"),
    ],
    "resolution-trends": [("Date", "date"), ("Resolved", "value")],
    "volume-trends": [("Date", "date"), ("Created", "value")],
    "created-vs-resolved": [
        ("Date", "date"), ("Created", "created"), ("Resolved", "resolved"), ("Net", "net"),
    ],
    "resolution-time-by-priority": [
        ("Priority", "priority"), ("Resolved", "resolved_count"),
        ("Avg (h)", "avg_hours"), ("Min (h)", "min_hours"), ("Max (h)", "max_hours"),
    ],
    "sla-breach-list": [
        ("Ticket", "ticket_number"), ("Summary", "summary"), ("SLA metric", "metric"),
        ("Priority", "priority"), ("Team", "team"), ("Due", "due_at"),
        ("Breached at", "breached_at"), ("Mins overdue", "minutes_overdue"),
    ],
    "backlog-aging": [("Age bucket", "label"), ("Open", "value")],
}


def _open_tickets_rows(data):
    rows = [["Open total", data.get("total", 0)]]
    for r in data.get("by_project", []):
        rows.append([f"  {r.get('project__key') or '—'}", r.get("n", 0)])
    return ["Metric", "Value"], rows


def _sla_rows(data):
    pct = data.get("compliance_pct")
    return ["Metric", "Value"], [
        ["Compliance %", "—" if pct is None else f"{pct}%"],
        ["Total measured", data.get("total", 0)],
        ["Met", data.get("met", 0)],
        ["Breached", data.get("breached", 0)],
    ]


def _table_rows(key, data):
    """(header_list, list-of-row-lists) for a list-shaped report."""
    cols = _TABLES[key]
    headers = [h for h, _ in cols]
    rows = []
    for item in data if isinstance(data, list) else []:
        rows.append([_fmt(item.get(k)) for _, k in cols])
    return headers, rows


def _fmt(v):
    if v is None:
        return "—"
    return v


def _manifest_rows(data):
    """(headers, rows) for a report that ships its own ``{columns, rows}`` manifest
    (e.g. ``ticket-data``, whose columns are dynamic — system + SLA + custom fields)."""
    cols = (data or {}).get("columns", [])
    headers = [c["label"] for c in cols]
    rows = [[_fmt(r.get(c["key"])) for c in cols] for r in (data or {}).get("rows", [])]
    return headers, rows


def _section(key, data):
    """Normalise any report's data into (title, headers, rows)."""
    title = REPORT_TITLES.get(key, key)
    if isinstance(data, dict) and "columns" in data and "rows" in data:
        headers, rows = _manifest_rows(data)
    elif key == "open-tickets":
        headers, rows = _open_tickets_rows(data)
    elif key == "sla-compliance":
        headers, rows = _sla_rows(data)
    elif key in _TABLES:
        headers, rows = _table_rows(key, data)
    else:  # unknown / future report — best-effort flatten
        headers, rows = ["Key", "Value"], [[k, _fmt(v)] for k, v in (data or {}).items()]
    return title, headers, rows


def _sheet_title(key):
    """openpyxl sheet titles are ≤31 chars and forbid []:*?/\\."""
    t = REPORT_TITLES.get(key, key)
    for ch in "[]:*?/\\":
        t = t.replace(ch, " ")
    return t[:31]


def build_workbook(report_data, *, scope_label, generated_at):
    """Return ``.xlsx`` bytes: one sheet per report + a cover sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="4F46E5")
    title_font = Font(bold=True, size=14)

    cover = wb.active
    cover.title = "Summary"
    cover["A1"] = "ITSM report pack"
    cover["A1"].font = title_font
    cover["A3"] = "Scope"
    cover["B3"] = scope_label
    cover["A4"] = "Generated"
    cover["B4"] = generated_at.strftime("%Y-%m-%d %H:%M")
    cover["A5"] = "Reports"
    cover["B5"] = ", ".join(REPORT_TITLES.get(k, k) for k in report_data)
    cover.column_dimensions["A"].width = 16
    cover.column_dimensions["B"].width = 80

    for key, data in report_data.items():
        title, headers, rows = _section(key, data)
        ws = wb.create_sheet(_sheet_title(key))
        ws["A1"] = title
        ws["A1"].font = title_font
        hr = 3
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=hr, column=c, value=h)
            cell.font, cell.fill = head_font, head_fill
            cell.alignment = Alignment(horizontal="left")
        for r, row in enumerate(rows, start=hr + 1):
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=val)
        if not rows:
            ws.cell(row=hr + 1, column=1, value="No data.")
        for c in range(1, len(headers) + 1):
            width = max(
                [len(str(headers[c - 1]))] + [len(str(row[c - 1])) for row in rows]
            ) + 4
            ws.column_dimensions[ws.cell(row=hr, column=c).column_letter].width = min(width, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_csv(report_key, data):
    """Return UTF-8 (BOM) CSV bytes for a single report — one table.

    The BOM makes Excel open unicode correctly on a double-click.
    """
    _title, headers, rows = _section(report_key, data)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(["" if v is None else v for v in row])
    return buf.getvalue().encode("utf-8-sig")
